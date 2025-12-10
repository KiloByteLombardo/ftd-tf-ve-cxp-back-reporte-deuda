"""
Programa para procesar archivos de Excel: Ordenes de Compra y Tasa
"""
import pandas as pd
import os
from typing import Tuple, Optional, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Imports para Google Sheets
GSPREAD_AVAILABLE = False
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    pass  # gspread no está disponible, pero no es crítico para el funcionamiento básico

# Columnas esperadas en el archivo de Ordenes de Compra
COLUMNAS_ESPERADAS = [
    'NUMERO_OC',
    'PROVEEDOR',
    'SUCURSAL',
    'DIVISA',
    'PRICE_OVERRIDE',
    'IMPORTE',
    'IMPORTE_RECIBIDO',
    'IMPORTE_ASOCIADO',
    'FECHA_ORDEN',
    'UNIDAD_MEDIDA',
    'DESCRIPCION',
    'CUENTA_CARGO',
    'SOLICITANTE',
    'ESTADO_CIERRE',
    'APROBADOR',
    'FECHA_CIERRE'
]


def detectar_cabezales(df: pd.DataFrame, num_filas_revisar: int = 5) -> int:
    """
    Detecta automáticamente la fila donde están los cabezales.
    Busca la fila que contiene todas las columnas esperadas.
    
    Args:
        df: DataFrame sin procesar
        num_filas_revisar: Número de filas a revisar desde el inicio
        
    Returns:
        Índice de la fila que contiene los cabezales, o 0 si no se encuentra
    """
    for i in range(min(num_filas_revisar, len(df))):
        # Normalizar los nombres de las columnas de la fila actual
        fila_actual = df.iloc[i].astype(str).str.upper().str.strip()
        valores_fila = set(fila_actual.values)
        
        # Normalizar las columnas esperadas
        columnas_esperadas_normalizadas = {col.upper().strip() for col in COLUMNAS_ESPERADAS}
        
        # Verificar si esta fila contiene todas las columnas esperadas
        if columnas_esperadas_normalizadas.issubset(valores_fila):
            print(f"Cabezales detectados en la fila {i + 1}")
            return i
    
    print("⚠ No se detectaron cabezales automáticamente, usando la primera fila")
    return 0


def verificar_columnas(df: pd.DataFrame) -> bool:
    """
    Verifica que el DataFrame tenga exactamente las 16 columnas esperadas.
    
    Args:
        df: DataFrame a verificar
        
    Returns:
        True si tiene las columnas correctas, False en caso contrario
    """
    columnas_df = set(col.upper().strip() for col in df.columns)
    columnas_esperadas = set(col.upper().strip() for col in COLUMNAS_ESPERADAS)
    
    if columnas_df == columnas_esperadas:
        print("✓ Todas las columnas están presentes y correctas")
        return True
    else:
        faltantes = columnas_esperadas - columnas_df
        extras = columnas_df - columnas_esperadas
        
        if faltantes:
            print(f"✗ Columnas faltantes: {faltantes}")
        if extras:
            print(f"⚠ Columnas extras encontradas: {extras}")
        
        return False


def leer_ordenes_compra(ruta_archivo: str) -> pd.DataFrame:
    """
    Lee el archivo de Ordenes de Compra con detección automática de cabezales.
    
    Args:
        ruta_archivo: Ruta al archivo Excel de Ordenes de Compra
        
    Returns:
        DataFrame con los datos procesados
    """
    print(f"Leyendo archivo de Ordenes de Compra: {ruta_archivo}")
    
    # Leer el archivo sin cabezales primero para detectar dónde están
    df_raw = pd.read_excel(ruta_archivo, header=None)
    
    # Detectar la fila de cabezales
    fila_cabezales = detectar_cabezales(df_raw)
    
    # Leer el archivo con los cabezales correctos
    df = pd.read_excel(ruta_archivo, header=fila_cabezales)
    
    # Normalizar nombres de columnas (mayúsculas y sin espacios extra)
    df.columns = df.columns.str.upper().str.strip()
    
    # Verificar que tenga las columnas correctas
    if not verificar_columnas(df):
        raise ValueError("El archivo no tiene las columnas esperadas")
    
    print(f"✓ Archivo leído correctamente. Filas: {len(df)}")
    return df


def detectar_cabezales_tasa(df: pd.DataFrame, num_filas_revisar: int = 10) -> int:
    """
    Detecta automáticamente la fila donde están los cabezales en el archivo de Tasa.
    Busca la fila que contiene las columnas esperadas de tasa.
    
    Args:
        df: DataFrame sin procesar del archivo de Tasa
        num_filas_revisar: Número de filas a revisar desde el inicio
        
    Returns:
        Índice de la fila que contiene los cabezales, o 0 si no se encuentra
    """
    # Columnas esperadas en el archivo de Tasa (basado en la imagen)
    columnas_esperadas_tasa = ['FECHA', 'VES/USD', 'VES/EUR', 'COP/USD', 'EUR/USD', 'COP/VES', 'VES/COF']
    
    for i in range(min(num_filas_revisar, len(df))):
        # Normalizar los valores de la fila actual
        fila_actual = df.iloc[i].astype(str).str.upper().str.strip()
        valores_fila = set(fila_actual.values)
        
        # Normalizar las columnas esperadas
        columnas_esperadas_normalizadas = {col.upper().strip() for col in columnas_esperadas_tasa}
        
        # Verificar si esta fila contiene al menos algunas de las columnas esperadas
        # (no necesariamente todas, porque puede haber columnas adicionales)
        coincidencias = columnas_esperadas_normalizadas.intersection(valores_fila)
        
        # Si encontramos al menos 3 columnas esperadas, probablemente es la fila de cabezales
        if len(coincidencias) >= 3:
            print(f"Cabezales de Tasa detectados en la fila {i + 1}")
            print(f"  Columnas encontradas: {coincidencias}")
            return i
    
    print("⚠ No se detectaron cabezales de Tasa automáticamente, usando la primera fila")
    return 0


def leer_tasa(ruta_archivo: str) -> pd.DataFrame:
    """
    Lee el archivo de Tasa con detección automática de cabezales.
    
    Args:
        ruta_archivo: Ruta al archivo Excel de Tasa
        
    Returns:
        DataFrame con los datos de tasa
    """
    print(f"Leyendo archivo de Tasa: {ruta_archivo}")
    
    # Verificar si hay múltiples hojas
    try:
        xl_file = pd.ExcelFile(ruta_archivo)
        print(f"  Hojas encontradas: {xl_file.sheet_names}")
        # Usar la primera hoja (o puedes especificar cuál usar)
        sheet_name = xl_file.sheet_names[1]
        print(f"  Leyendo hoja: {sheet_name}")
    except Exception as e:
        print(f"⚠ Error al leer información de hojas: {e}")
        sheet_name = 0
    
    # Leer el archivo sin cabezales primero para detectar dónde están
    # Usar nrows=None para leer todas las filas
    df_raw = pd.read_excel(ruta_archivo, sheet_name=sheet_name, header=None, engine='openpyxl')
    
    print(f"  Filas leídas sin cabezales: {len(df_raw)}")
    
    # Detectar la fila de cabezales
    fila_cabezales = detectar_cabezales_tasa(df_raw)
    
    # Leer el archivo con los cabezales correctos
    # Usar engine='openpyxl' y asegurarse de leer todas las filas
    try:
        df = pd.read_excel(
            ruta_archivo, 
            sheet_name=sheet_name,
            header=fila_cabezales,
            engine='openpyxl'
        )
        print(f"  Filas leídas con cabezales: {len(df)}")
    except Exception as e:
        print(f"⚠ Error al leer con cabezales detectados: {e}")
        # Intentar sin cabezales como fallback
        df = pd.read_excel(
            ruta_archivo, 
            sheet_name=sheet_name,
            header=None,
            engine='openpyxl'
        )
        fila_cabezales = 0
    
    # Normalizar nombres de columnas
    df.columns = df.columns.str.upper().str.strip()
    
    # Limpiar solo filas completamente vacías (todas las columnas son NaN)
    # Pero mantener filas que tengan al menos un valor
    filas_antes = len(df)
    df = df.dropna(how='all')
    filas_despues = len(df)
    
    if filas_antes != filas_despues:
        print(f"  Filas completamente vacías eliminadas: {filas_antes - filas_despues}")
    
    print(f"✓ Archivo de Tasa leído correctamente. Filas: {len(df)}, Columnas: {len(df.columns)}")
    print(f"  Columnas encontradas: {list(df.columns)}")
    
    # Verificar si hay más datos después de filas vacías
    # Si el número de filas es menor al esperado, intentar leer de nuevo sin límites
    if len(df) < 1000:  # Si hay menos de 1000 filas, puede que falten datos
        print(f"⚠ Advertencia: Solo se leyeron {len(df)} filas. Verificando si hay más datos...")
        # Intentar leer todas las filas sin procesar
        try:
            df_completo = pd.read_excel(
                ruta_archivo,
                sheet_name=sheet_name,
                header=None,
                engine='openpyxl'
            )
            print(f"  Total de filas en archivo (sin procesar): {len(df_completo)}")
            
            # Si hay más filas, leer de nuevo desde el cabezal hasta el final
            if len(df_completo) > len(df) + fila_cabezales + 10:
                print(f"  Re-leyendo archivo completo desde cabezal...")
                df = pd.read_excel(
                    ruta_archivo,
                    sheet_name=sheet_name,
                    header=fila_cabezales,
                    engine='openpyxl'
                )
                df.columns = df.columns.str.upper().str.strip()
                df = df.dropna(how='all')
                print(f"  Filas después de re-lectura: {len(df)}")
        except Exception as e:
            print(f"⚠ Error al verificar filas adicionales: {e}")
    
    return df


def agregar_columna_tasa(df_ordenes: pd.DataFrame, df_tasa: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega la columna TASA al DataFrame de ordenes usando un VLOOKUP basado en FECHA_ORDEN y DIVISA.
    Busca la tasa correspondiente según la divisa (VES/USD, COP/USD, EUR/USD).
    Si no encuentra la fecha exacta, usa la del día anterior o la fila anterior.
    
    Args:
        df_ordenes: DataFrame de Ordenes de Compra
        df_tasa: DataFrame de Tasas
        
    Returns:
        DataFrame de ordenes con la columna TASA agregada
    """
    print("Agregando columna TASA...")
    
    # Buscar la columna de fecha en el DataFrame de tasa
    columna_fecha_tasa = None
    posibles_fechas = ['FECHA', 'DATE', 'FECHA_CAMBIO', 'FECHA_TASA']
    for col in df_tasa.columns:
        if any(pf in col.upper() for pf in posibles_fechas):
            columna_fecha_tasa = col
            break
    
    if not columna_fecha_tasa:
        columna_fecha_tasa = df_tasa.columns[0]
        print(f"⚠ No se encontró columna de fecha explícita, usando: {columna_fecha_tasa}")
    
    # Buscar columnas de tasa según divisa
    columnas_tasa = {}
    tasas_buscadas = {
        'VES': ['VES/USD', 'VES_USD'],
        'COP': ['COP/USD', 'COP_USD'],
        'EUR': ['EUR/USD', 'EUR_USD']
    }
    
    for divisa, patrones in tasas_buscadas.items():
        for col in df_tasa.columns:
            col_upper = col.upper()
            if any(patron.upper() in col_upper for patron in patrones):
                columnas_tasa[divisa] = col
                print(f"  Columna de tasa {divisa} encontrada: {col}")
                break
    
    if not columnas_tasa:
        # Fallback: buscar cualquier columna que contenga "USD" o "TASA"
        for col in df_tasa.columns:
            col_upper = col.upper()
            if 'USD' in col_upper or 'TASA' in col_upper or 'CAMBIO' in col_upper:
                columnas_tasa['DEFAULT'] = col
                print(f"⚠ Usando columna de tasa por defecto: {col}")
                break
    
    print(f"  Columna fecha en tasa: {columna_fecha_tasa}")
    
    # Preparar DataFrame de tasas ordenado por fecha
    df_tasa_clean = df_tasa.copy()
    df_tasa_clean[columna_fecha_tasa] = pd.to_datetime(df_tasa_clean[columna_fecha_tasa], errors='coerce')
    df_tasa_clean = df_tasa_clean.sort_values(by=columna_fecha_tasa)
    df_tasa_clean = df_tasa_clean.dropna(subset=[columna_fecha_tasa])
    
    # Crear DataFrame de lookup con todas las tasas
    tasas_lookup = []
    for _, row in df_tasa_clean.iterrows():
        fecha = row[columna_fecha_tasa]
        if pd.isna(fecha):
            continue
        
        fecha_str = fecha.strftime('%Y-%m-%d') if isinstance(fecha, pd.Timestamp) else str(fecha)
        tasas_row = {'fecha': fecha, 'fecha_str': fecha_str}
        
        # Agregar cada tipo de tasa
        for divisa, col_tasa in columnas_tasa.items():
            if col_tasa in row.index:
                tasas_row[f'tasa_{divisa}'] = row[col_tasa]
        
        tasas_lookup.append(tasas_row)
    
    # Crear DataFrame de lookup para búsqueda eficiente
    df_lookup = pd.DataFrame(tasas_lookup)
    
    # Función para buscar tasa con fallback a fecha anterior
    def buscar_tasa_con_fallback(fecha_orden, divisa):
        if pd.isna(fecha_orden):
            return None
        
        try:
            fecha_dt = pd.to_datetime(fecha_orden)
            fecha_str = fecha_dt.strftime('%Y-%m-%d')
            
            # Normalizar divisa
            divisa_upper = str(divisa).upper().strip() if not pd.isna(divisa) else ""
            
            # Determinar qué columna de tasa usar
            col_tasa_key = None
            if divisa_upper == 'VES':
                col_tasa_key = 'tasa_VES'
            elif divisa_upper == 'COP':
                col_tasa_key = 'tasa_COP'
            elif divisa_upper == 'EUR':
                col_tasa_key = 'tasa_EUR'
            else:
                # Para USD o cualquier otra divisa, usar la primera tasa disponible como referencia
                # Prioridad: VES/USD > COP/USD > EUR/USD > DEFAULT
                if 'tasa_VES' in df_lookup.columns:
                    col_tasa_key = 'tasa_VES'
                elif 'tasa_COP' in df_lookup.columns:
                    col_tasa_key = 'tasa_COP'
                elif 'tasa_EUR' in df_lookup.columns:
                    col_tasa_key = 'tasa_EUR'
                elif 'tasa_DEFAULT' in df_lookup.columns:
                    col_tasa_key = 'tasa_DEFAULT'
                else:
                    # Si no hay ninguna tasa disponible, retornar None
                    return None
            
            # Buscar fecha exacta
            fila_exacta = df_lookup[df_lookup['fecha_str'] == fecha_str]
            if not fila_exacta.empty and col_tasa_key in fila_exacta.columns:
                tasa = fila_exacta.iloc[0][col_tasa_key]
                if not pd.isna(tasa):
                    return tasa
            
            # Si no se encuentra, buscar la fecha anterior más cercana
            filas_anteriores = df_lookup[df_lookup['fecha'] <= fecha_dt]
            if not filas_anteriores.empty:
                # Tomar la última fila (más cercana a la fecha buscada)
                fila_anterior = filas_anteriores.iloc[-1]
                if col_tasa_key in fila_anterior.index:
                    tasa = fila_anterior[col_tasa_key]
                    if not pd.isna(tasa):
                        print(f"  ⚠ Tasa no encontrada para fecha {fecha_str}, usando fecha anterior: {fila_anterior['fecha_str']}")
                        return tasa
            
            return None
        except Exception as e:
            print(f"⚠ Error al buscar tasa para fecha {fecha_orden}, divisa {divisa}: {e}")
            return None
    
    # Agregar columna TASA al DataFrame de ordenes
    df_ordenes_con_tasa = df_ordenes.copy()
    
    # Asegurar que existe la columna DIVISA
    if 'DIVISA' not in df_ordenes_con_tasa.columns:
        print("⚠ Advertencia: No existe columna DIVISA, no se puede determinar la tasa correcta")
        df_ordenes_con_tasa['TASA'] = None
        return df_ordenes_con_tasa
    
    # Aplicar búsqueda de tasa
    df_ordenes_con_tasa['TASA'] = df_ordenes_con_tasa.apply(
        lambda row: buscar_tasa_con_fallback(row['FECHA_ORDEN'], row['DIVISA']), 
        axis=1
    )
    
    # Contar cuántas tasas se encontraron
    tasas_encontradas = df_ordenes_con_tasa['TASA'].notna().sum()
    print(f"✓ Columna TASA agregada. Tasas encontradas: {tasas_encontradas}/{len(df_ordenes_con_tasa)}")
    
    return df_ordenes_con_tasa


def aplicar_estilos_excel(ruta_archivo: str):
    """
    Aplica estilos a las hojas del archivo Excel.
    Cabezales: fondo azul, texto blanco.
    
    Args:
        ruta_archivo: Ruta al archivo Excel
    """
    print("Aplicando estilos al archivo Excel...")
    
    wb = load_workbook(ruta_archivo)
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Azul
    header_font = Font(color="FFFFFF", bold=True, size=11)  # Blanco, negrita
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar estilos a cada hoja
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Aplicar estilos a la primera fila (cabezales)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ajustar ancho con un poco de margen
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Aplicar bordes a todas las celdas con datos
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = border
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
    
    wb.save(ruta_archivo)
    print("✓ Estilos aplicados exitosamente")


def agregar_ano_fiscal(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega la columna AÑO FISCAL basada en FECHA_ORDEN.
    
    Lógica:
    - Si el mes > 8 (septiembre-diciembre): "20XX-20XX+1"
    - Si el mes <= 8 (enero-agosto): "20XX-1-20XX"
    
    Args:
        df: DataFrame de Ordenes de Compra
        
    Returns:
        DataFrame con la columna AÑO FISCAL agregada
    """
    print("Agregando columna AÑO FISCAL...")
    
    if 'FECHA_ORDEN' not in df.columns:
        raise ValueError("La columna FECHA_ORDEN no existe en el DataFrame")
    
    df_con_ano_fiscal = df.copy()
    
    def calcular_ano_fiscal(fecha_orden):
        """
        Calcula el año fiscal basado en la fecha de orden.
        """
        if pd.isna(fecha_orden):
            return None
        
        try:
            # Convertir a datetime si no lo es
            fecha_dt = pd.to_datetime(fecha_orden)
            año = fecha_dt.year
            mes = fecha_dt.month
            
            # Si el mes es mayor a 8 (septiembre, octubre, noviembre, diciembre)
            if mes > 8:
                # Año fiscal: "20XX-20XX+1"
                año_siguiente = año + 1
                return f"{año}-{año_siguiente}"
            else:
                # Si el mes es <= 8 (enero-agosto)
                # Año fiscal: "20XX-1-20XX"
                año_anterior = año - 1
                return f"{año_anterior}-{año}"
        except Exception as e:
            print(f"⚠ Error al calcular año fiscal para fecha {fecha_orden}: {e}")
            return None
    
    df_con_ano_fiscal['AÑO FISCAL'] = df_con_ano_fiscal['FECHA_ORDEN'].apply(calcular_ano_fiscal)
    
    # Contar cuántos años fiscales se calcularon
    anos_fiscales_calculados = df_con_ano_fiscal['AÑO FISCAL'].notna().sum()
    print(f"✓ Columna AÑO FISCAL agregada. Años fiscales calculados: {anos_fiscales_calculados}/{len(df_con_ano_fiscal)}")
    
    return df_con_ano_fiscal


def agregar_montos_oc(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega las columnas MONTO OC y MONTO OC USD.
    
    Lógica:
    - MONTO OC = PRICE_OVERRIDE * IMPORTE
    - MONTO OC USD = 
      * Si DIVISA == "USD": MONTO OC (ya está en USD, no se divide)
      * Si DIVISA == "VES": MONTO OC / TASA (VES/USD)
      * Si DIVISA == "COP": MONTO OC / TASA (COP/USD)
      * Si DIVISA == "EUR": MONTO OC / TASA (EUR/USD)
      * Todo lo que no sea USD se convierte a USD dividiendo por la tasa correspondiente
    
    Args:
        df: DataFrame de Ordenes de Compra (debe tener TASA ya agregada)
        
    Returns:
        DataFrame con las columnas MONTO OC y MONTO OC USD agregadas
    """
    print("Agregando columnas MONTO OC y MONTO OC USD...")
    
    # Verificar que existan las columnas necesarias
    columnas_requeridas = ['PRICE_OVERRIDE', 'IMPORTE', 'DIVISA', 'TASA']
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        raise ValueError(f"Faltan las siguientes columnas: {columnas_faltantes}")
    
    df_con_montos = df.copy()
    
    # Calcular MONTO OC = PRICE_OVERRIDE * IMPORTE
    def calcular_monto_oc(row):
        try:
            price_override = row['PRICE_OVERRIDE']
            importe = row['IMPORTE']
            
            # Manejar valores nulos
            if pd.isna(price_override) or pd.isna(importe):
                return None
            
            # Convertir a numérico si es necesario
            price_override = pd.to_numeric(price_override, errors='coerce')
            importe = pd.to_numeric(importe, errors='coerce')
            
            if pd.isna(price_override) or pd.isna(importe):
                return None
            
            return price_override * importe
        except Exception as e:
            print(f"⚠ Error al calcular MONTO OC: {e}")
            return None
    
    df_con_montos['MONTO OC'] = df_con_montos.apply(calcular_monto_oc, axis=1)
    
    # Calcular MONTO OC USD
    def calcular_monto_oc_usd(row):
        try:
            monto_oc = row['MONTO OC']
            divisa = row['DIVISA']
            tasa = row['TASA']
            
            # Si MONTO OC es nulo, retornar nulo
            if pd.isna(monto_oc):
                return None
            
            # Convertir divisa a string y normalizar
            divisa_str = str(divisa).upper().strip() if not pd.isna(divisa) else ""
            
            # Si la divisa es USD, no necesita conversión (ya está en USD)
            if divisa_str == 'USD':
                return monto_oc
            
            # Si no es USD (VES, COP, EUR), convertir a USD dividiendo por la tasa
            if divisa_str in ['VES', 'COP', 'EUR']:
                if pd.isna(tasa) or tasa == 0:
                    print(f"⚠ Tasa nula o cero para divisa {divisa_str}, no se puede convertir a USD")
                    return None
                
                # Convertir tasa a numérico
                tasa_num = pd.to_numeric(tasa, errors='coerce')
                if pd.isna(tasa_num) or tasa_num == 0:
                    print(f"⚠ Tasa inválida para divisa {divisa_str}, no se puede convertir a USD")
                    return None
                
                return monto_oc / tasa_num
            else:
                # Para otras divisas no reconocidas, intentar usar la tasa si existe
                if not pd.isna(tasa) and tasa != 0:
                    tasa_num = pd.to_numeric(tasa, errors='coerce')
                    if not pd.isna(tasa_num) and tasa_num != 0:
                        return monto_oc / tasa_num
                
                print(f"⚠ Divisa {divisa_str} no reconocida y sin tasa, no se puede convertir a USD")
                return None
        except Exception as e:
            print(f"⚠ Error al calcular MONTO OC USD: {e}")
            return None
    
    df_con_montos['MONTO OC USD'] = df_con_montos.apply(calcular_monto_oc_usd, axis=1)
    
    # Contar cuántos montos se calcularon
    montos_oc_calculados = df_con_montos['MONTO OC'].notna().sum()
    montos_oc_usd_calculados = df_con_montos['MONTO OC USD'].notna().sum()
    
    print(f"✓ Columnas MONTO OC y MONTO OC USD agregadas.")
    print(f"  MONTO OC calculados: {montos_oc_calculados}/{len(df_con_montos)}")
    print(f"  MONTO OC USD calculados: {montos_oc_usd_calculados}/{len(df_con_montos)}")
    
    return df_con_montos


def agregar_montos_oc_asociado(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega las columnas MONTO OC ASOCIADO y MONTO OC ASOCIADO USD.
    
    Lógica:
    - MONTO OC ASOCIADO = IMPORTE_ASOCIADO * PRICE_OVERRIDE
    - MONTO OC ASOCIADO USD = 
      * Si DIVISA == "USD": MONTO OC ASOCIADO (ya está en USD, no se divide)
      * Si DIVISA == "VES": MONTO OC ASOCIADO / TASA (VES/USD)
      * Si DIVISA == "COP": MONTO OC ASOCIADO / TASA (COP/USD)
      * Si DIVISA == "EUR": MONTO OC ASOCIADO / TASA (EUR/USD)
      * Todo lo que no sea USD se convierte a USD dividiendo por la tasa correspondiente
    
    Args:
        df: DataFrame de Ordenes de Compra (debe tener TASA ya agregada)
        
    Returns:
        DataFrame con las columnas MONTO OC ASOCIADO y MONTO OC ASOCIADO USD agregadas
    """
    print("Agregando columnas MONTO OC ASOCIADO y MONTO OC ASOCIADO USD...")
    
    # Verificar que existan las columnas necesarias
    columnas_requeridas = ['PRICE_OVERRIDE', 'IMPORTE_ASOCIADO', 'DIVISA', 'TASA']
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        raise ValueError(f"Faltan las siguientes columnas: {columnas_faltantes}")
    
    df_con_montos_asociado = df.copy()
    
    # Calcular MONTO OC ASOCIADO = IMPORTE_ASOCIADO * PRICE_OVERRIDE
    def calcular_monto_oc_asociado(row):
        try:
            price_override = row['PRICE_OVERRIDE']
            importe_asociado = row['IMPORTE_ASOCIADO']
            
            # Manejar valores nulos
            if pd.isna(price_override) or pd.isna(importe_asociado):
                return None
            
            # Convertir a numérico si es necesario
            price_override = pd.to_numeric(price_override, errors='coerce')
            importe_asociado = pd.to_numeric(importe_asociado, errors='coerce')
            
            if pd.isna(price_override) or pd.isna(importe_asociado):
                return None
            
            return importe_asociado * price_override
        except Exception as e:
            print(f"⚠ Error al calcular MONTO OC ASOCIADO: {e}")
            return None
    
    df_con_montos_asociado['MONTO OC ASOCIADO'] = df_con_montos_asociado.apply(calcular_monto_oc_asociado, axis=1)
    
    # Calcular MONTO OC ASOCIADO USD
    def calcular_monto_oc_asociado_usd(row):
        try:
            monto_oc_asociado = row['MONTO OC ASOCIADO']
            divisa = row['DIVISA']
            tasa = row['TASA']
            
            # Si MONTO OC ASOCIADO es nulo, retornar nulo
            if pd.isna(monto_oc_asociado):
                return None
            
            # Convertir divisa a string y normalizar
            divisa_str = str(divisa).upper().strip() if not pd.isna(divisa) else ""
            
            # Si la divisa es USD, no necesita conversión (ya está en USD)
            if divisa_str == 'USD':
                return monto_oc_asociado
            
            # Si no es USD (VES, COP, EUR), convertir a USD dividiendo por la tasa
            if divisa_str in ['VES', 'COP', 'EUR']:
                if pd.isna(tasa) or tasa == 0:
                    print(f"⚠ Tasa nula o cero para divisa {divisa_str}, no se puede convertir a USD")
                    return None
                
                # Convertir tasa a numérico
                tasa_num = pd.to_numeric(tasa, errors='coerce')
                if pd.isna(tasa_num) or tasa_num == 0:
                    print(f"⚠ Tasa inválida para divisa {divisa_str}, no se puede convertir a USD")
                    return None
                
                return monto_oc_asociado / tasa_num
            else:
                # Para otras divisas no reconocidas, intentar usar la tasa si existe
                if not pd.isna(tasa) and tasa != 0:
                    tasa_num = pd.to_numeric(tasa, errors='coerce')
                    if not pd.isna(tasa_num) and tasa_num != 0:
                        return monto_oc_asociado / tasa_num
                
                print(f"⚠ Divisa {divisa_str} no reconocida y sin tasa, no se puede convertir a USD")
                return None
        except Exception as e:
            print(f"⚠ Error al calcular MONTO OC ASOCIADO USD: {e}")
            return None
    
    df_con_montos_asociado['MONTO OC ASOCIADO USD'] = df_con_montos_asociado.apply(calcular_monto_oc_asociado_usd, axis=1)
    
    # Contar cuántos montos se calcularon
    montos_oc_asociado_calculados = df_con_montos_asociado['MONTO OC ASOCIADO'].notna().sum()
    montos_oc_asociado_usd_calculados = df_con_montos_asociado['MONTO OC ASOCIADO USD'].notna().sum()
    
    print(f"✓ Columnas MONTO OC ASOCIADO y MONTO OC ASOCIADO USD agregadas.")
    print(f"  MONTO OC ASOCIADO calculados: {montos_oc_asociado_calculados}/{len(df_con_montos_asociado)}")
    print(f"  MONTO OC ASOCIADO USD calculados: {montos_oc_asociado_usd_calculados}/{len(df_con_montos_asociado)}")
    
    return df_con_montos_asociado


def agregar_monto_real_deuda(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega la columna MONTO REAL DEUDA.
    
    Lógica:
    - MONTO REAL DEUDA = MONTO OC USD - MONTO OC ASOCIADO USD
    
    Args:
        df: DataFrame de Ordenes de Compra (debe tener MONTO OC USD y MONTO OC ASOCIADO USD)
        
    Returns:
        DataFrame con la columna MONTO REAL DEUDA agregada
    """
    print("Agregando columna MONTO REAL DEUDA...")
    
    # Verificar que existan las columnas necesarias
    columnas_requeridas = ['MONTO OC USD', 'MONTO OC ASOCIADO USD']
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        raise ValueError(f"Faltan las siguientes columnas: {columnas_faltantes}")
    
    df_con_monto_real = df.copy()
    
    # Calcular MONTO REAL DEUDA = MONTO OC USD - MONTO OC ASOCIADO USD
    def calcular_monto_real_deuda(row):
        try:
            monto_oc_usd = row['MONTO OC USD']
            monto_oc_asociado_usd = row['MONTO OC ASOCIADO USD']
            
            # Manejar valores nulos
            # Si ambos son nulos, retornar nulo
            if pd.isna(monto_oc_usd) and pd.isna(monto_oc_asociado_usd):
                return None
            
            # Convertir a numérico si es necesario
            monto_oc_usd = pd.to_numeric(monto_oc_usd, errors='coerce') if not pd.isna(monto_oc_usd) else 0
            monto_oc_asociado_usd = pd.to_numeric(monto_oc_asociado_usd, errors='coerce') if not pd.isna(monto_oc_asociado_usd) else 0
            
            # Si alguno no se pudo convertir, usar 0
            if pd.isna(monto_oc_usd):
                monto_oc_usd = 0
            if pd.isna(monto_oc_asociado_usd):
                monto_oc_asociado_usd = 0
            
            return monto_oc_usd - monto_oc_asociado_usd
        except Exception as e:
            print(f"⚠ Error al calcular MONTO REAL DEUDA: {e}")
            return None
    
    df_con_monto_real['MONTO REAL DEUDA'] = df_con_monto_real.apply(calcular_monto_real_deuda, axis=1)
    
    # Contar cuántos montos se calcularon
    montos_reales_calculados = df_con_monto_real['MONTO REAL DEUDA'].notna().sum()
    
    print(f"✓ Columna MONTO REAL DEUDA agregada.")
    print(f"  MONTO REAL DEUDA calculados: {montos_reales_calculados}/{len(df_con_monto_real)}")
    
    return df_con_monto_real


def preparar_dataframe_bigquery(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara el DataFrame para BigQuery mapeando columnas y agregando timestamp.
    
    Args:
        df: DataFrame con todas las columnas procesadas
        
    Returns:
        DataFrame preparado para BigQuery con nombres de columnas mapeados
    """
    print("Preparando DataFrame para BigQuery...")
    
    # Mapeo de columnas del DataFrame a nombres de BigQuery
    mapeo_columnas = {
        'NUMERO_OC': 'vzla_deuda_orden_compra',
        'PROVEEDOR': 'vzla_deuda_proveedor',
        'SUCURSAL': 'vzla_deuda_sucursal',
        'DIVISA': 'vzla_deuda_divisa',
        'PRICE_OVERRIDE': 'vzla_deuda_price_override',
        'IMPORTE': 'vzla_deuda_importe',
        'IMPORTE_RECIBIDO': 'vzla_deuda_importe_recibido',
        'IMPORTE_ASOCIADO': 'vzla_deuda_importe_asociado',
        'FECHA_ORDEN': 'vzla_deuda_fecha_orden',
        'UNIDAD_MEDIDA': 'vzla_deuda_unidad_medida',
        'DESCRIPCION': 'vzla_deuda_descripcion',
        'CUENTA_CARGO': 'vzla_deuda_cuenta_cargo',
        'SOLICITANTE': 'vzla_deuda_solicitante',
        'ESTADO_CIERRE': 'vzla_deuda_estado_cierre',
        'APROBADOR': 'vzla_deuda_aprobador',
        'FECHA_CIERRE': 'vzla_deuda_fecha_cierre',
        'AÑO FISCAL': 'vzla_deuda_fiscal_year',
        'TASA': 'vzla_deuda_tasa',
        'AREA': 'vzla_deuda_area',
        'MONTO OC': 'vzla_deuda_monto_oc',
        'MONTO OC USD': 'vzla_deuda_monto_oc_usd',
        'MONTO OC ASOCIADO': 'vzla_deuda_monto_oc_asociado',
        'MONTO OC ASOCIADO USD': 'vzla_deuda_monto_oc_asociado_usd',
        'MONTO REAL DEUDA': 'vzla_deuda_monto_real_deuda'
    }
    
    # Columnas que deben ser STRING en BigQuery
    columnas_string = [
        'vzla_deuda_orden_compra',
        'vzla_deuda_proveedor',
        'vzla_deuda_sucursal',
        'vzla_deuda_divisa',
        'vzla_deuda_unidad_medida',
        'vzla_deuda_descripcion',
        'vzla_deuda_cuenta_cargo',
        'vzla_deuda_solicitante',
        'vzla_deuda_estado_cierre',
        'vzla_deuda_aprobador',
        'vzla_deuda_fiscal_year',
        'vzla_deuda_area'
    ]
    
    df_bq = df.copy()
    
    # Renombrar columnas según el mapeo
    df_bq = df_bq.rename(columns=mapeo_columnas)
    
    # Agregar timestamp
    timestamp_actual = datetime.now()
    df_bq['vzla_deuda_timestamp'] = timestamp_actual
    
    # Seleccionar solo las columnas que existen en el DataFrame
    columnas_finales = [col for col in mapeo_columnas.values() if col in df_bq.columns]
    columnas_finales.append('vzla_deuda_timestamp')
    
    df_bq = df_bq[columnas_finales]
    
    # Convertir columnas STRING a string explícitamente (para evitar errores con int64)
    for col in columnas_string:
        if col in df_bq.columns:
            # Convertir a string, manejando valores nulos
            df_bq[col] = df_bq[col].astype(str).replace('nan', None).replace('None', None)
            # Si el valor es 'None' o 'nan', dejarlo como None
            df_bq[col] = df_bq[col].apply(lambda x: None if pd.isna(x) or str(x).lower() in ['none', 'nan', ''] else str(x))
    
    print(f"✓ DataFrame preparado para BigQuery. Columnas: {len(columnas_finales)}")
    return df_bq


def subir_a_bigquery(df: pd.DataFrame, project_id: str, dataset_id: str, table_id: str) -> bool:
    """
    Sube el DataFrame a BigQuery.
    
    Args:
        df: DataFrame preparado para BigQuery
        project_id: ID del proyecto de GCP
        dataset_id: ID del dataset de BigQuery
        table_id: ID de la tabla de BigQuery
        
    Returns:
        True si se subió exitosamente, False en caso contrario
    """
    print(f"Subiendo datos a BigQuery: {project_id}.{dataset_id}.{table_id}")
    
    try:
        from google.cloud import bigquery
        
        client = bigquery.Client(project=project_id)
        table_ref = client.dataset(dataset_id).table(table_id)
        
        # Convertir fechas a formato correcto
        df_upload = df.copy()
        
        # Imprimir tipos de datos antes de la conversión
        print("  Tipos de datos del DataFrame antes de conversión:")
        for col in df_upload.columns:
            print(f"    {col}: {df_upload[col].dtype}")
        
        # Convertir columnas de fecha a datetime si es necesario
        fecha_columns = ['vzla_deuda_fecha_orden', 'vzla_deuda_fecha_cierre']
        for col in fecha_columns:
            if col in df_upload.columns:
                df_upload[col] = pd.to_datetime(df_upload[col], errors='coerce')
        
        # Asegurar que timestamp sea datetime
        if 'vzla_deuda_timestamp' in df_upload.columns:
            df_upload['vzla_deuda_timestamp'] = pd.to_datetime(df_upload['vzla_deuda_timestamp'])
        
        # Asegurar que todas las columnas STRING sean realmente string
        columnas_string = [
            'vzla_deuda_orden_compra', 'vzla_deuda_proveedor', 'vzla_deuda_sucursal',
            'vzla_deuda_divisa', 'vzla_deuda_unidad_medida', 'vzla_deuda_descripcion',
            'vzla_deuda_cuenta_cargo', 'vzla_deuda_solicitante', 'vzla_deuda_estado_cierre',
            'vzla_deuda_aprobador', 'vzla_deuda_fiscal_year', 'vzla_deuda_area'
        ]
        
        for col in columnas_string:
            if col in df_upload.columns:
                # Convertir a string, manejando valores nulos
                # Primero convertir a string, luego reemplazar 'nan' y 'None' con None
                df_upload[col] = df_upload[col].astype(str)
                df_upload[col] = df_upload[col].replace(['nan', 'None', '<NA>', 'NaT'], None)
                # Si el valor es None o NaN, dejarlo como None
                df_upload[col] = df_upload[col].apply(lambda x: None if pd.isna(x) or str(x).lower() in ['none', 'nan', '', '<na>', 'nat'] else str(x))
        
        # Imprimir tipos de datos después de la conversión
        print("  Tipos de datos del DataFrame después de conversión:")
        for col in df_upload.columns:
            print(f"    {col}: {df_upload[col].dtype}")
        
        # Configuración del job - usar WRITE_APPEND para agregar datos
        job_config = bigquery.LoadJobConfig(
            write_disposition="WRITE_APPEND",
        )
        
        # Cargar DataFrame directamente a BigQuery
        job = client.load_table_from_dataframe(df_upload, table_ref, job_config=job_config)
        job.result()  # Esperar a que termine el job
        
        print(f"✓ Datos subidos a BigQuery exitosamente. Filas: {len(df_upload)}")
        return True
        
    except Exception as e:
        print(f"✗ Error al subir a BigQuery: {e}")
        import traceback
        traceback.print_exc()
        return False


def subir_excel_a_cloud_storage(ruta_archivo: str, bucket_name: str, nombre_archivo: Optional[str] = None) -> Optional[str]:
    """
    Sube el archivo Excel a Cloud Storage y retorna la URL pública.
    
    Args:
        ruta_archivo: Ruta local al archivo Excel
        bucket_name: Nombre del bucket de Cloud Storage
        nombre_archivo: Nombre del archivo en Cloud Storage (opcional)
        
    Returns:
        URL pública del archivo en Cloud Storage, o None si hay error
    """
    print(f"Subiendo archivo a Cloud Storage: {bucket_name}")
    
    try:
        from google.cloud import storage
        
        client = storage.Client()
        bucket = client.bucket(bucket_name)
        
        # Generar nombre del archivo si no se proporciona
        if not nombre_archivo:
            nombre_archivo = os.path.basename(ruta_archivo)
        
        # Subir el archivo
        blob = bucket.blob(nombre_archivo)
        blob.upload_from_filename(ruta_archivo)
        
        print(f"  Archivo subido como: {nombre_archivo}")
        
        # Hacer el archivo público y obtener la URL
        try:
            blob.make_public()
            url_publica = blob.public_url
            print(f"  URL obtenida de blob.public_url: {url_publica}")
        except Exception as e:
            print(f"  ⚠ Error al hacer público el blob: {e}")
            url_publica = None
        
        # Verificar que la URL se generó correctamente
        if not url_publica or url_publica == '' or url_publica is None:
            # Si public_url no funciona, construir la URL manualmente
            url_publica = f"https://storage.googleapis.com/{bucket_name}/{nombre_archivo}"
            print(f"  URL construida manualmente: {url_publica}")
        
        print(f"✓ Archivo subido a Cloud Storage exitosamente")
        print(f"  URL pública final: {url_publica}")
        print(f"  Bucket: {bucket_name}")
        print(f"  Blob: {nombre_archivo}")
        
        return url_publica
        
    except Exception as e:
        print(f"✗ Error al subir a Cloud Storage: {e}")
        import traceback
        traceback.print_exc()
        return None


def leer_areas_desde_sheets(spreadsheet_id: str, credentials_path: Optional[str] = None) -> Dict[str, str]:
    """
    Lee la tabla de SOLICITANTE y AREA desde Google Sheets.
    
    Args:
        spreadsheet_id: ID del Google Sheet
        credentials_path: Ruta opcional al archivo de credenciales JSON
        
    Returns:
        Diccionario con SOLICITANTE como clave y AREA como valor
    """
    print("Leyendo áreas desde Google Sheets...")
    
    if not GSPREAD_AVAILABLE:
        print("✗ Error: gspread no está disponible")
        return {}
    
    try:
        # Obtener credenciales
        # Prioridad 1: Archivo de credenciales proporcionado explícitamente (desarrollo local)
        if credentials_path and os.path.exists(credentials_path):
            print(f"  Usando archivo de credenciales proporcionado: {credentials_path}")
            creds = Credentials.from_service_account_file(
                credentials_path,
                scopes=['https://www.googleapis.com/auth/spreadsheets.readonly']
            )
            gc = gspread.authorize(creds)
        else:
            # Prioridad 2: Variable de entorno GOOGLE_APPLICATION_CREDENTIALS (desarrollo local)
            creds_path = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
            if creds_path and os.path.exists(creds_path):
                print(f"  Usando GOOGLE_APPLICATION_CREDENTIALS desde: {creds_path}")
                creds = Credentials.from_service_account_file(
                    creds_path,
                    scopes=['https://www.googleapis.com/auth/spreadsheets.readonly']
                )
                gc = gspread.authorize(creds)
            else:
                # Prioridad 3: Application Default Credentials (Cloud Run)
                print("  Usando Application Default Credentials (ADC) - Cloud Run")
                try:
                    from google.auth import default
                    creds, project = default(scopes=['https://www.googleapis.com/auth/spreadsheets.readonly'])
                    gc = gspread.authorize(creds)
                    print("  ✓ Credenciales ADC obtenidas exitosamente")
                except Exception as adc_error:
                    print(f"✗ Error: No se pudieron obtener credenciales de Google Sheets")
                    print(f"  Error ADC: {adc_error}")
                    print("  En Cloud Run, asegúrate de que el service account tenga acceso al Google Sheet")
                    return {}
        
        # Conectar con Google Sheets
        sheet = gc.open_by_key(spreadsheet_id)
        
        # Obtener la primera hoja (o puedes especificar el nombre)
        worksheet = sheet.sheet1
        
        # Leer todos los datos
        datos = worksheet.get_all_values()
        
        if len(datos) < 2:
            print("⚠ Advertencia: El Google Sheet tiene menos de 2 filas (solo cabezales)")
            return {}
        
        # Buscar las columnas SOLICITANTE y AREA en la primera fila
        headers = [str(h).upper().strip() for h in datos[0]]
        
        col_solicitante_idx = None
        col_area_idx = None
        
        for idx, header in enumerate(headers):
            if 'SOLICITANTE' in header:
                col_solicitante_idx = idx
            if 'AREA' in header:
                col_area_idx = idx
        
        if col_solicitante_idx is None or col_area_idx is None:
            print(f"⚠ Advertencia: No se encontraron las columnas SOLICITANTE y AREA")
            print(f"  Columnas encontradas: {headers}")
            return {}
        
        print(f"  Columna SOLICITANTE encontrada en índice: {col_solicitante_idx}")
        print(f"  Columna AREA encontrada en índice: {col_area_idx}")
        
        # Crear diccionario de SOLICITANTE -> AREA
        areas_dict = {}
        for row in datos[1:]:  # Saltar la primera fila (cabezales)
            if len(row) > max(col_solicitante_idx, col_area_idx):
                solicitante = str(row[col_solicitante_idx]).strip() if row[col_solicitante_idx] else ""
                area = str(row[col_area_idx]).strip() if row[col_area_idx] else ""
                
                if solicitante and area:
                    areas_dict[solicitante.upper()] = area
        
        print(f"✓ Áreas leídas desde Google Sheets: {len(areas_dict)} registros")
        return areas_dict
        
    except Exception as e:
        print(f"✗ Error al leer Google Sheets: {e}")
        import traceback
        traceback.print_exc()
        return {}


def agregar_columna_area(df: pd.DataFrame, spreadsheet_id: Optional[str] = None, credentials_path: Optional[str] = None) -> pd.DataFrame:
    """
    Agrega la columna AREA haciendo pareo con Google Sheets usando SOLICITANTE.
    
    Args:
        df: DataFrame de Ordenes de Compra
        spreadsheet_id: ID del Google Sheet (opcional, puede venir de variable de entorno)
        credentials_path: Ruta opcional al archivo de credenciales JSON
        
    Returns:
        DataFrame con la columna AREA agregada
    """
    print("Agregando columna AREA...")
    
    if 'SOLICITANTE' not in df.columns:
        raise ValueError("La columna SOLICITANTE no existe en el DataFrame")
    
    # Obtener spreadsheet_id de variable de entorno si no se proporciona
    if not spreadsheet_id:
        spreadsheet_id = os.getenv('GOOGLE_SHEETS_SPREADSHEET_ID', "15JAM-L4wTWSAs1wUrHFWBpYmgTDTGys2m36CcaI4UjU")
    
    if not spreadsheet_id:
        print("⚠ Advertencia: No se proporcionó GOOGLE_SHEETS_SPREADSHEET_ID, no se puede agregar columna AREA")
        df_con_area = df.copy()
        df_con_area['AREA'] = None
        return df_con_area
    
    # Leer áreas desde Google Sheets
    areas_dict = leer_areas_desde_sheets(spreadsheet_id, credentials_path)
    
    if not areas_dict:
        print("⚠ Advertencia: No se pudieron leer áreas desde Google Sheets")
        df_con_area = df.copy()
        df_con_area['AREA'] = None
        return df_con_area
    
    df_con_area = df.copy()
    
    # Hacer el pareo usando SOLICITANTE
    def buscar_area(solicitante):
        if pd.isna(solicitante):
            return None
        
        try:
            solicitante_str = str(solicitante).upper().strip()
            return areas_dict.get(solicitante_str)
        except Exception as e:
            print(f"⚠ Error al buscar área para solicitante {solicitante}: {e}")
            return None
    
    df_con_area['AREA'] = df_con_area['SOLICITANTE'].apply(buscar_area)
    
    # Contar cuántas áreas se encontraron
    areas_encontradas = df_con_area['AREA'].notna().sum()
    print(f"✓ Columna AREA agregada. Áreas encontradas: {areas_encontradas}/{len(df_con_area)}")
    
    return df_con_area


def filtrar_cerrados(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra las filas donde ESTADO_CIERRE == "CERRADO".
    
    Args:
        df: DataFrame de Ordenes de Compra
        
    Returns:
        DataFrame filtrado sin las filas con ESTADO_CIERRE == "CERRADO"
    """
    if 'ESTADO_CIERRE' not in df.columns:
        raise ValueError("La columna ESTADO_CIERRE no existe en el DataFrame")
    
    filas_antes = len(df)
    df_filtrado = df[df['ESTADO_CIERRE'].astype(str).str.upper().str.strip() != 'CERRADO'].copy()
    filas_despues = len(df_filtrado)
    filas_eliminadas = filas_antes - filas_despues
    
    print(f"✓ Filtrado completado. Filas eliminadas: {filas_eliminadas}, Filas restantes: {filas_despues}")
    
    return df_filtrado


def procesar_archivos(
    ruta_ordenes: str,
    ruta_tasa: str,
    ruta_salida: Optional[str] = None
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Procesa ambos archivos Excel y genera el resultado.
    
    Args:
        ruta_ordenes: Ruta al archivo de Ordenes de Compra
        ruta_tasa: Ruta al archivo de Tasa
        ruta_salida: Ruta opcional para guardar el resultado
        
    Returns:
        Tupla con (DataFrame de ordenes procesadas, DataFrame de tasa)
    """
    # Leer archivos
    df_ordenes = leer_ordenes_compra(ruta_ordenes)
    df_tasa = leer_tasa(ruta_tasa)
    
    # Filtrar filas con ESTADO_CIERRE == "CERRADO"
    df_ordenes_filtrado = filtrar_cerrados(df_ordenes)
    
    # Agregar columna AÑO FISCAL
    df_ordenes_con_ano_fiscal = agregar_ano_fiscal(df_ordenes_filtrado)
    
    # Agregar columna TASA usando VLOOKUP por fecha
    df_ordenes_con_tasa = agregar_columna_tasa(df_ordenes_con_ano_fiscal, df_tasa)
    
    # Agregar columna AREA desde Google Sheets
    # En Cloud Run, no se necesita credentials_path (usa ADC automáticamente)
    # En desarrollo local, se puede pasar credentials_path si es necesario
    credentials_path = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
    spreadsheet_id = os.getenv('GOOGLE_SHEETS_SPREADSHEET_ID')
    df_ordenes_con_area = agregar_columna_area(
        df_ordenes_con_tasa, 
        spreadsheet_id, 
        credentials_path if credentials_path and os.path.exists(credentials_path) else None
    )
    
    # Agregar columnas MONTO OC y MONTO OC USD
    df_ordenes_con_montos = agregar_montos_oc(df_ordenes_con_area)
    
    # Agregar columnas MONTO OC ASOCIADO y MONTO OC ASOCIADO USD
    df_ordenes_con_montos_asociado = agregar_montos_oc_asociado(df_ordenes_con_montos)
    
    # Agregar columna MONTO REAL DEUDA
    df_ordenes_final = agregar_monto_real_deuda(df_ordenes_con_montos_asociado)
    
    # Guardar resultado si se especifica ruta de salida
    if ruta_salida:
        print(f"Guardando resultado en: {ruta_salida}")
        
        # Crear ExcelWriter para múltiples hojas
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            # Hoja 1: Ordenes de Compra procesadas
            df_ordenes_final.to_excel(writer, sheet_name='Ordenes de Compra', index=False)
            
            # Hoja 2: Tasas
            df_tasa.to_excel(writer, sheet_name='Tasa', index=False)
        
        # Aplicar estilos al archivo
        aplicar_estilos_excel(ruta_salida)
        print("✓ Archivo guardado exitosamente")
    
    return df_ordenes_final, df_tasa


if __name__ == "__main__":
    # Ejemplo de uso
    import sys
    
    ruta_ordenes = "D:/Users/andres.moreno/Downloads/Deuda Farmatodo VEN/Reporte Ordenes de Compras VZLA (1).xlsx"
    ruta_tasa = "D:/Users/andres.moreno/Downloads/Deuda Farmatodo VEN/Tasa de Cambio (ME).xlsx"
    
    # Generar nombre de archivo con fecha actual (formato: día_mes_año)
    fecha_actual = datetime.now()
    fecha_formato = f"{fecha_actual.day}_{fecha_actual.month}_{fecha_actual.year}"
    directorio_salida = "D:/Users/andres.moreno/Documents/Deuda Farmatodo VZLA/deuda_vzla/resultados"
    nombre_archivo = f"resultado_deuda_{fecha_formato}.xlsx"
    ruta_salida = os.path.join(directorio_salida, nombre_archivo)
    
    try:
        df_ordenes, df_tasa = procesar_archivos(ruta_ordenes, ruta_tasa, ruta_salida)
        print(f"✓ Procesamiento completado. Filas procesadas: {len(df_ordenes)}")
        print(f"✓ Archivo guardado en: {ruta_salida}")
    except Exception as e:
        print(f"✗ Error al procesar archivos: {e}")
        sys.exit(1)

